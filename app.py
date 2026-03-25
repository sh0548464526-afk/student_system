
import os
from flask import Flask, render_template, request, redirect, session, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from openpyxl import Workbook

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "secret")

db_url = os.getenv("DATABASE_URL", "sqlite:///db.sqlite3")
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ========= MODELS =========
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True)
    password = db.Column(db.String(200))

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100))
    tz = db.Column(db.String(20))

class Seder(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100))
    start_time = db.Column(db.String(5))
    amount = db.Column(db.Float)
    deduction = db.Column(db.Float)
    late_minutes = db.Column(db.Integer)

class Day(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    day = db.Column(db.String(10))
    marked = db.Column(db.Boolean)

class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey("student.id"))
    s1 = db.Column(db.String(5))
    s2 = db.Column(db.String(5))
    s3 = db.Column(db.String(5))
    total = db.Column(db.Float)
    student = db.relationship("Student")

# ========= AUTH =========
@app.route("/", methods=["GET","POST"])
def login():
    if request.method == "POST":
        user = User.query.filter_by(username=request.form["username"]).first()
        if user and check_password_hash(user.password, request.form["password"]):
            session["user"] = user.username
            return redirect("/students")
    return render_template("login.html")

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect("/")
        return f(*args, **kwargs)
    return decorated

# ========= STUDENTS =========
@app.route("/students", methods=["GET","POST"])
@login_required
def students():
    if request.method == "POST":
        sid = request.form.get("id")
        if sid:
            s = Student.query.get(sid)
            s.name = request.form["name"]
            s.tz = request.form["tz"]
        else:
            s = Student(name=request.form["name"], tz=request.form["tz"])
            db.session.add(s)
        db.session.commit()
        return redirect("/students")
    return render_template("students.html", data=Student.query.all())

# ========= SEDARIM =========
@app.route("/sedarim", methods=["GET","POST"])
@login_required
def sedarim():
    if request.method == "POST":
        sid = request.form.get("id")
        if sid:
            s = Seder.query.get(sid)
            s.name = request.form["name"]
            s.start_time = request.form["start_time"]
            s.amount = float(request.form["amount"])
            s.deduction = float(request.form["deduction"])
            s.late_minutes = int(request.form["late_minutes"])
        else:
            s = Seder(
                name=request.form["name"],
                start_time=request.form["start_time"],
                amount=float(request.form["amount"]),
                deduction=float(request.form["deduction"]),
                late_minutes=int(request.form["late_minutes"])
            )
            db.session.add(s)
        db.session.commit()
        return redirect("/sedarim")
    return render_template("sedarim.html", data=Seder.query.all())

# ========= DAYS =========
@app.route("/days", methods=["GET","POST"])
@login_required
def days():
    if request.method == "POST":
        did = request.form.get("id")
        marked = "marked" in request.form
        if did:
            d = Day.query.get(did)
            d.day = request.form["day"]
            d.marked = marked
        else:
            d = Day(day=request.form["day"], marked=marked)
            db.session.add(d)
        db.session.commit()
        return redirect("/days")
    return render_template("days.html", data=Day.query.all())

# ========= ATTENDANCE =========
def calc_total(att):
    sedarim = Seder.query.order_by(Seder.id).all()
    times = [att.s1, att.s2, att.s3]
    total = 0
    for i, t in enumerate(times):
        if i >= len(sedarim) or not t:
            continue
        seder = sedarim[i]
        try:
            arrival = datetime.strptime(t, "%H:%M")
            start = datetime.strptime(seder.start_time, "%H:%M")
            diff = (arrival - start).total_seconds() / 60
            if diff <= 0:
                total += seder.amount
            else:
                units = diff // seder.late_minutes
                total += max(0, seder.amount - units * seder.deduction)
        except:
            pass
    return round(total,2)

@app.route("/attendance", methods=["GET","POST"])
@login_required
def attendance():
    students = Student.query.all()
    if request.method == "POST":
        att_id = request.form.get("id")
        student_id = int(request.form["student_id"])
        s1 = request.form["s1"]
        s2 = request.form["s2"]
        s3 = request.form["s3"]
        if att_id:
            att = Attendance.query.get(att_id)
            att.student_id = student_id
            att.s1 = s1
            att.s2 = s2
            att.s3 = s3
        else:
            att = Attendance(student_id=student_id, s1=s1, s2=s2, s3=s3)
            db.session.add(att)
        att.total = calc_total(att)
        db.session.commit()
        return redirect("/attendance")
    return render_template("attendance.html", data=Attendance.query.all(), students=students)

# ========= EXPORT =========
@app.route("/export")
@login_required
def export():
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["שם","תז"])
    for s in Student.query.all():
        ws.append([s.name, s.tz])

    ws2 = wb.create_sheet("Sedarim")
    ws2.append(["שם סדר","שעת התחלה","סכום","קיזוז","X דקות"])
    for s in Seder.query.all():
        ws2.append([s.name, s.start_time, s.amount, s.deduction, s.late_minutes])

    ws3 = wb.create_sheet("Days")
    ws3.append(["יום","סימון"])
    for d in Day.query.all():
        ws3.append([d.day, d.marked])

    ws4 = wb.create_sheet("Attendance")
    ws4.append(["שם","תז","סדר1","סדר2","סדר3","סכום"])
    for i, att in enumerate(Attendance.query.all(), start=2):
        ws4[f"A{i}"] = att.student.name
        ws4[f"B{i}"] = att.student.tz
        ws4[f"C{i}"] = att.s1
        ws4[f"D{i}"] = att.s2
        ws4[f"E{i}"] = att.s3
        ws4[f"F{i}"] = f"=SUM(C{i}:E{i})"

    filename = f"עדכון נכון ל-{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    wb.save(filename)
    return send_file(filename, as_attachment=True)

# ========= INIT =========
with app.app_context():
    db.create_all()
    if not User.query.first():
        db.session.add(User(username="admin", password=generate_password_hash("1234")))
        db.session.commit()

# ========= RUN =========
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)
